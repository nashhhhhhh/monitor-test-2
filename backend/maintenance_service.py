from __future__ import annotations

import calendar
import copy
import json
import os
import re
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd


MAINTENANCE_WORKBOOK_PATH = Path(
    os.environ.get(
        "MAINTENANCE_WORKBOOK_PATH",
        r"C:\Users\merri\Downloads\utility maintenance\FO-FC11-011 Preventive Maintenance Schedule 2025J 2025.xlsx",
    )
)

EQUIPMENT_MAINTENANCE_WORKBOOK_PATH = Path(
    os.environ.get(
        "EQUIPMENT_MAINTENANCE_WORKBOOK_PATH",
        r"C:\Users\merri\Downloads\FO-FC11-011 Preventive Maintenance Schedule 2025 stage 2.xlsx",
    )
)
EQUIPMENT_MAINTENANCE_COPY_PATH = Path(__file__).resolve().parent.parent / "data" / "equipment_maintenance_source_copy.xlsx"
EQUIPMENT_MAINTENANCE_CACHE_PATH = Path(__file__).resolve().parent.parent / "data" / "equipment_maintenance_cache.json"
EQUIPMENT_MAINTENANCE_CACHE_VERSION = 4
ADDITIONAL_STEPS_LABEL = "Additional steps required beyond the normal checklist"

_MAINTENANCE_CACHE = {}

MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_NAME_TO_NUMBER = {label.lower(): index + 1 for index, label in enumerate(MONTH_LABELS)}
WEEK_ORDER = {
    "first": 1,
    "second": 2,
    "third": 3,
    "fourth": 4,
    "last": -1,
}
WEEK_LABELS = {
    "first": "1st week",
    "second": "2nd week",
    "third": "3rd week",
    "fourth": "4th week",
    "last": "Last week",
    "every_week": "Every week",
}

LOCATION_EXACT_MAP = {
    "อาคารบอยเลอร์": "Boiler Room",
    "ห้องปั๊มลม": "Air Compressor Room",
    "โรงบำบัดน้ำดี": "Water Treatment Plant",
    "โรงบำบัดน้ำเสีย": "Wastewater Treatment Plant",
    "บ่อบาดาลบ่อ 1": "Deep Well 1",
    "บ่อบาดาลบ่อ 2": "Deep Well 2",
    "บ่อบาดาลบ่อ1": "Deep Well 1",
    "บ่อบาดาลบ่อ2": "Deep Well 2",
}

LOCATION_REPLACEMENTS = [
    ("ระบบเติมอากาศ", "Air Intake System"),
    ("ระบบดูดอากาศ", "Exhaust System"),
    ("ห้อง Cooking", "Cooking Room"),
    ("ห้องล้างฝั่งดิบ", "Raw Wash Area"),
    ("ฝั่งดิบ", "Raw Side"),
    ("ห้อง", "Room "),
]

EQUIPMENT_RISK_MAP = {
    "cooking": "Low Risk",
    "prep": "Low Risk",
    "preparation": "Low Risk",
    "cooking and prep": "Low Risk",
    "outgoing": "Medium Risk",
    "out going": "Medium Risk",
    "cartoning": "Medium Risk",
    "inbound": "Medium Risk",
    "packing": "Medium Risk",
    "high risk": "High Risk",
}


def get_file_signature(path: Path):
    try:
        stat = path.stat()
    except OSError:
        return None
    return (stat.st_mtime_ns, stat.st_size)


def load_json_cache(path: Path):
    try:
        if not path.exists():
            return None
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def write_json_cache(path: Path, payload):
    try:
        path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception:
        return


def resolve_equipment_workbook_path() -> Path:
    if EQUIPMENT_MAINTENANCE_COPY_PATH.exists():
        return EQUIPMENT_MAINTENANCE_COPY_PATH
    return EQUIPMENT_MAINTENANCE_WORKBOOK_PATH


def clean_text(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def contains_thai(text: str | None) -> bool:
    return bool(text and re.search(r"[\u0E00-\u0E7F]", text))


def normalize_asset_code(value) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return text.upper()


def normalize_location_display(location_raw: str | None) -> str | None:
    cleaned = clean_text(location_raw)
    if not cleaned:
        return None

    if cleaned in LOCATION_EXACT_MAP:
        return LOCATION_EXACT_MAP[cleaned]

    translated = cleaned
    for source, replacement in LOCATION_REPLACEMENTS:
        translated = translated.replace(source, replacement)
    translated = re.sub(r"\s+", " ", translated).strip(" -")

    return translated if not contains_thai(translated) else cleaned


def worksheet_is_utility_schedule(frame: pd.DataFrame) -> bool:
    for row_index in range(min(6, len(frame.index))):
        row_values = {
            clean_text(value)
            for value in frame.iloc[row_index].tolist()
            if clean_text(value)
        }
        if "Machine Code" in row_values and "Machine Name" in row_values:
            return True
    return False


def worksheet_has_maintenance_header(frame: pd.DataFrame) -> bool:
    return worksheet_is_utility_schedule(frame)


def worksheet_has_equipment_header(worksheet) -> bool:
    max_row = min(8, worksheet.max_row)
    max_col = min(6, worksheet.max_column)
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        row_values = {
            clean_text(value)
            for value in row
            if clean_text(value)
        }
        if "Machine Code" in row_values and "Machine Name" in row_values:
            return True
    return False


def find_maintenance_header_row(frame: pd.DataFrame) -> int | None:
    for row_index in range(min(8, len(frame.index))):
        row_values = {
            clean_text(value)
            for value in frame.iloc[row_index].tolist()
            if clean_text(value)
        }
        if "Machine Code" in row_values and "Machine Name" in row_values:
            return row_index
    return None


def parse_month_token(value) -> int | None:
    text = clean_text(value)
    if not text:
        return None

    lowered = text.lower()
    for label, number in MONTH_NAME_TO_NUMBER.items():
        if lowered.startswith(label):
            return number

    match = re.search(r"\b(1[0-2]|0?[1-9])\b", lowered)
    if match and any(month in lowered for month in ["month", "mth"]):
        return int(match.group(1))

    return None


def parse_week_token(value) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    lowered = text.lower()
    compact = re.sub(r"[^a-z0-9]+", "", lowered)

    if "every week" in lowered or compact in {"everyweek", "weekly"}:
        return "every_week"
    if "last" in lowered or compact in {"week5", "w5", "5"}:
        return "last"
    if any(token in lowered for token in ["1st", "first"]) or compact in {"1", "wk1", "w1", "week1"}:
        return "first"
    if any(token in lowered for token in ["2nd", "second"]) or compact in {"2", "wk2", "w2", "week2"}:
        return "second"
    if any(token in lowered for token in ["3rd", "third"]) or compact in {"3", "wk3", "w3", "week3"}:
        return "third"
    if any(token in lowered for token in ["4th", "fourth"]) or compact in {"4", "wk4", "w4", "week4"}:
        return "fourth"

    return None


def is_schedule_marker(value) -> bool:
    text = clean_text(value)
    if not text:
        return False

    lowered = text.lower()
    if lowered in {"-", "--", "n/a", "na", "0", "no"}:
        return False

    return True


def canonicalize_equipment_category(value: str | None) -> str:
    cleaned = clean_text(value) or "General"
    normalized = re.sub(r"[^a-z0-9]+", " ", cleaned.lower()).strip()

    if "cooking" in normalized and "prep" in normalized:
        return "Cooking and Prep"
    if "cooking" in normalized:
        return "Cooking"
    if "prep" in normalized:
        return "Prep"
    if "outgoing" in normalized:
        return "Outgoing"
    if "cartoning" in normalized:
        return "Cartoning"
    if "inbound" in normalized:
        return "Inbound"
    if "packing" in normalized:
        return "Packing"

    return cleaned


def get_equipment_risk_level(category: str | None) -> str:
    cleaned = canonicalize_equipment_category(category)
    normalized = re.sub(r"[^a-z0-9]+", " ", cleaned.lower()).strip()
    return EQUIPMENT_RISK_MAP.get(normalized, "Medium")


def normalize_equipment_risk_label(value: str | None) -> str | None:
    text = clean_text(value)
    if not text:
        return None

    lowered = text.lower()
    if "high" in lowered:
        return "High Risk"
    if "low" in lowered:
        return "Low Risk"
    if "medium" in lowered:
        return "Medium Risk"
    return None


def classify_equipment_subcategory(asset_name: str | None, raw_area: str | None, risk_level: str | None) -> str:
    combined = " ".join(filter(None, [clean_text(asset_name), clean_text(raw_area)])).lower()

    preparation_keywords = [
        "pre-cooking",
        "sort spices",
        "sort rice",
        "chx prep",
        "veg prep",
        "fish prep",
        "meatball making",
        "meat grinder",
        "can opener",
        "dicing machine",
    ]
    if any(keyword in combined for keyword in preparation_keywords):
        return "Preparation"

    if "packing" in combined:
        return "Packing"
    if "out going" in combined or "outgoing" in combined:
        return "Out going"
    if "cartoning" in combined:
        return "Cartoning"
    if "inbound" in combined:
        return "Inbound"
    if "cooking" in combined:
        return "Cooking"

    if risk_level == "Low Risk":
        return "Cooking"
    if risk_level == "Medium Risk":
        return "Out going"
    if risk_level == "High Risk":
        return "Packing"
    return "Production Equipment"


def derive_equipment_category(subcategory: str | None, risk_level: str | None) -> str:
    cleaned_subcategory = clean_text(subcategory) or "Production Equipment"
    if clean_text(risk_level) == "High Risk":
        return "High Risk"

    normalized = cleaned_subcategory.lower()
    if normalized in {"packing", "out going", "outgoing", "cartoning", "inbound"}:
        return "Medium Risk"
    return "Low Risk"


def get_equipment_week_columns(worksheet) -> list[tuple[int, int]]:
    week_columns = []
    header_row = 4
    for column_index in range(5, worksheet.max_column + 1):
        value = worksheet.cell(header_row, column_index).value
        try:
            week_number = int(float(value))
        except (TypeError, ValueError):
            continue
        if 1 <= week_number <= 53:
            week_columns.append((column_index, week_number))
    return week_columns


def has_schedule_fill(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return False

    color = None
    try:
        color = fill.fgColor.rgb or fill.start_color.rgb
    except Exception:
        color = None

    # The workbook uses filled week cells to mark PM schedule slots.
    # Ignore empty/default fills, but allow any explicit colored fill.
    if color in {None, "00000000", "000000", "FFFFFFFF", "00FFFFFF"}:
        return False
    return True


def get_cell_color_rgb(color) -> str | None:
    if color is None:
        return None
    try:
        rgb = color.rgb
        if isinstance(rgb, str):
            return rgb.upper()
    except Exception:
        pass
    return None


def has_red_indicator(cell) -> bool:
    fill = getattr(cell, "fill", None)
    fill_rgb = get_cell_color_rgb(getattr(fill, "fgColor", None)) or get_cell_color_rgb(getattr(fill, "start_color", None))
    if fill_rgb and fill_rgb.endswith("FF0000"):
        return True

    font = getattr(cell, "font", None)
    font_rgb = get_cell_color_rgb(getattr(font, "color", None))
    if font_rgb and font_rgb.endswith("FF0000"):
        return True

    return False


def is_schedule_cell_marked(cell) -> bool:
    return has_schedule_fill(cell) or has_red_indicator(cell)


def iso_week_start(year: int, week_number: int) -> date:
    return date.fromisocalendar(year, week_number, 1)


def build_schedule_column_templates(frame: pd.DataFrame, header_row: int, skip_columns: set[int]) -> list[dict]:
    templates = []
    current_month = None

    for column_index in range(frame.shape[1]):
        if column_index in skip_columns:
            continue

        month_number = None
        week_key = None
        for row_index in range(header_row + 1):
            cell_value = frame.iat[row_index, column_index]
            parsed_month = parse_month_token(cell_value)
            if parsed_month:
                month_number = parsed_month

            parsed_week = parse_week_token(cell_value)
            if parsed_week:
                week_key = parsed_week

        if month_number:
            current_month = month_number

        if current_month and week_key:
            templates.append(
                {
                    "column_index": column_index,
                    "month": current_month,
                    "target_week": week_key,
                }
            )

    return templates


def extract_utility_inspection_templates(worksheet, frame: pd.DataFrame, sheet_name: str):
    header_row = find_maintenance_header_row(frame)
    if header_row is None:
        return {}

    templates = build_schedule_column_templates(frame, header_row, {0, 1, 2, 3})
    if not templates:
        return {}

    inspection_map = {}
    data_start_row = header_row + 2
    for row_offset, row in enumerate(frame.iloc[header_row + 1 :].itertuples(index=False), start=data_start_row):
        asset_code = normalize_asset_code(row[1] if len(row) > 1 else None)
        asset_name_raw = clean_text(row[2] if len(row) > 2 else None)

        if not asset_code or not asset_name_raw or not asset_code.startswith("UL-"):
            continue

        templates_for_asset = []
        for template in templates:
            column_number = int(template["column_index"]) + 1
            cell = worksheet.cell(row_offset, column_number)
            if not has_red_indicator(cell):
                continue
            templates_for_asset.append(
                {
                    "month": template["month"],
                    "target_week": template["target_week"],
                    "inspection_required": True,
                    "inspection_label": ADDITIONAL_STEPS_LABEL,
                }
            )

        if templates_for_asset:
            inspection_map[asset_code] = templates_for_asset

    return inspection_map


def extract_equipment_assets(worksheet, sheet_name: str, year: int):
    week_columns = get_equipment_week_columns(worksheet)
    if not week_columns:
        return []

    assets = []
    data_start_row = 5
    for row_index in range(data_start_row, worksheet.max_row + 1):
        asset_code = normalize_asset_code(worksheet.cell(row_index, 2).value)
        asset_name_raw = clean_text(worksheet.cell(row_index, 3).value)
        area_or_risk_raw = clean_text(worksheet.cell(row_index, 4).value)

        if not asset_code or not asset_name_raw:
            continue
        if asset_code.startswith("UL-"):
            continue

        due_templates = []
        for column_index, week_number in week_columns:
            cell = worksheet.cell(row_index, column_index)
            if not is_schedule_cell_marked(cell):
                continue
            due_templates.append(
                {
                    "week_number": week_number,
                    "inspection_required": has_red_indicator(cell),
                    "inspection_label": ADDITIONAL_STEPS_LABEL if has_red_indicator(cell) else None,
                }
            )

        if not due_templates:
            continue

        schedule_dates = []
        normalized_templates = []
        for template in sorted(due_templates, key=lambda item: item["week_number"]):
            week_number = template["week_number"]
            try:
                scheduled_date = iso_week_start(year, week_number)
            except ValueError:
                continue
            schedule_dates.append(scheduled_date)
            normalized_templates.append(
                {
                    "month": scheduled_date.month,
                    "target_week": "scheduled",
                    "scheduled_date": scheduled_date.isoformat(),
                    "inspection_required": bool(template.get("inspection_required")),
                    "inspection_label": template.get("inspection_label"),
                }
            )

        if not schedule_dates:
            continue

        first_date = schedule_dates[0]
        risk_level = normalize_equipment_risk_label(area_or_risk_raw)
        subcategory_value = classify_equipment_subcategory(asset_name_raw, area_or_risk_raw or sheet_name, risk_level)
        category_value = derive_equipment_category(subcategory_value, risk_level)
        assets.append(
            {
                "asset_code": asset_code,
                "asset_name": asset_name_raw,
                "asset_name_raw": asset_name_raw,
                "category": category_value,
                "subcategory": subcategory_value,
                "risk_level": risk_level,
                "location_raw": area_or_risk_raw,
                "location_display": subcategory_value,
                "location_detail": None,
                "source_sheet": sheet_name,
                "remarks": None,
                "scheduled_weeks": sorted({template["week_number"] for template in due_templates}),
                "schedule_dates": [scheduled_date.isoformat() for scheduled_date in schedule_dates],
                "due_templates": normalized_templates,
                "frequency_type": "scheduled",
                "frequency_value": len(schedule_dates),
                "start_month": first_date.month,
                "target_week": "scheduled",
                "due_months": sorted({scheduled_date.month for scheduled_date in schedule_dates}),
            }
        )

    return assets


def extract_schedule_assets(frame: pd.DataFrame, sheet_name: str):
    assets = []
    for _, row in frame.iloc[4:].iterrows():
        asset_code = normalize_asset_code(row.iloc[1] if len(row) > 1 else None)
        asset_name_raw = clean_text(row.iloc[2] if len(row) > 2 else None)
        location_raw = clean_text(row.iloc[3] if len(row) > 3 else None)

        if not asset_code or not asset_name_raw:
            continue
        if not asset_code.startswith("UL-"):
            continue

        assets.append(
            {
                "asset_code": asset_code,
                "asset_name": asset_name_raw,
                "asset_name_raw": asset_name_raw,
                "category": None,
                "subcategory": None,
                "location_raw": location_raw,
                "location_display": normalize_location_display(location_raw),
                "source_sheet": sheet_name,
                "remarks": None,
            }
        )

    return assets


def add_schedule_rule(rule_map, *, category, subcategory, frequency_type, frequency_value, start_month, target_week, codes):
    due_months = list(range(1, 13)) if frequency_type == "monthly" else [((start_month - 1) + step) % 12 + 1 for step in range(0, 12, frequency_value)]
    for code in codes:
        rule_map[code] = {
            "category": category,
            "subcategory": subcategory,
            "frequency_type": frequency_type,
            "frequency_value": frequency_value,
            "start_month": start_month,
            "target_week": target_week,
            "due_months": due_months,
        }


def build_schedule_rules():
    rules = {}

    add_schedule_rule(
        rules,
        category="Boiler",
        subcategory="Boiler",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="first",
        codes=["UL-HB-01", "UL-BL-01", "UL-BL-02"],
    )

    add_schedule_rule(
        rules,
        category="Air Compressor",
        subcategory="Air Compressor",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="first",
        codes=["UL-AC-01", "UL-AC-02", "UL-TN-01", "UL-TN-02"],
    )
    add_schedule_rule(
        rules,
        category="Air Compressor",
        subcategory="Air Dryer",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="second",
        codes=["UL-AD-01"],
    )

    add_schedule_rule(
        rules,
        category="MDB",
        subcategory="MDB",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=3,
        target_week="third",
        codes=["UL-MDB-01", "UL-MDB-02", "UL-MDB-03", "UL-MDB-04", "UL-MDB-05"],
    )

    add_schedule_rule(
        rules,
        category="Water Treatment",
        subcategory="Filtration",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="second",
        codes=["UL-SF-01", "UL-SF-02", "UL-CF-01", "UL-CF-02"],
    )
    add_schedule_rule(
        rules,
        category="Water Treatment",
        subcategory="Pump and Resin",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="third",
        codes=["UL-RS-01", "UL-RS-02", "UL-DP-01", "UL-DP-02", "UL-TP-01"],
    )
    add_schedule_rule(
        rules,
        category="Water Treatment",
        subcategory="Transfer Pump",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="fourth",
        codes=["UL-TP-02", "UL-TP-03", "UL-TP-04"],
    )
    add_schedule_rule(
        rules,
        category="Water Treatment",
        subcategory="Fire Pump",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=3,
        target_week="last",
        codes=["UL-FP-01"],
    )
    add_schedule_rule(
        rules,
        category="Water Treatment",
        subcategory="RO Water",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="second",
        codes=["UL-RO-01"],
    )

    add_schedule_rule(
        rules,
        category="Wastewater Treatment",
        subcategory="Air Blower",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="last",
        codes=["UL-AB-02", "UL-AB-03", "UL-AB-04", "UL-SP--01"],
    )
    add_schedule_rule(
        rules,
        category="Wastewater Treatment",
        subcategory="Submersible Pump",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=3,
        target_week="first",
        codes=["UL-SP--02", "UL-SP--03", "UL-SP--04", "UL-SP--05", "UL-SP--06", "UL-SP--07", "UL-SP--08"],
    )

    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Hood",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="third",
        codes=["UL-HD-01", "UL-HD-02", "UL-HD-03", "UL-HD-04"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Hood",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="last",
        codes=["UL-HD-05"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Intake",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=1,
        target_week="last",
        codes=["UL-IN-01", "UL-IN-02"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Intake",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="first",
        codes=["UL-IN-03", "UL-IN-04"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Exhaust",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="first",
        codes=["UL-EX-01", "UL-EX-02", "UL-EX-03"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Exhaust",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="second",
        codes=["UL-EX-04", "UL-EX-05", "UL-EX-06", "UL-EX-07", "UL-EX-08"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Exhaust",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=2,
        target_week="third",
        codes=["UL-EX-09"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="Transformer",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=3,
        target_week="second",
        codes=["UL-TR-01", "UL-TR-02", "UL-TR-03", "UL-TR-04", "UL-TR-05"],
    )
    add_schedule_rule(
        rules,
        category="Others",
        subcategory="UV Machine",
        frequency_type="monthly",
        frequency_value=1,
        start_month=1,
        target_week="every_week",
        codes=["UL-UV-01", "UL-UV-02", "UL-UV-03"],
    )

    add_schedule_rule(
        rules,
        category="LPG",
        subcategory="LPG",
        frequency_type="quarterly",
        frequency_value=3,
        start_month=3,
        target_week="last",
        codes=["UL-LPG-01", "UL-LPG-02", "UL-LPG-03", "UL-VP-01", "UL-VP-02"],
    )

    add_schedule_rule(
        rules,
        category="Laundry Room",
        subcategory="Laundry",
        frequency_type="monthly",
        frequency_value=1,
        start_month=1,
        target_week="last",
        codes=["UL-WH-01", "UL-WH-02", "UL-DR-01", "UL-DR-02"],
    )

    add_schedule_rule(
        rules,
        category="Others",
        subcategory="PP Filter",
        frequency_type="monthly",
        frequency_value=1,
        start_month=1,
        target_week="every_week",
        codes=["UL-PP-01", "UL-PP-02", "UL-PP-03", "UL-PP-04", "UL-PP-05", "UL-PP-06", "UL-PP-07"],
    )

    return rules


SCHEDULE_RULES = build_schedule_rules()

CANONICAL_ASSET_NAMES = {
    "UL-HB-01": "Hot Oil Boiler",
    "UL-BL-01": "Steam Boiler No.1",
    "UL-BL-02": "Steam Boiler No.2",
    "UL-AC-01": "Air Compressor No.1",
    "UL-AC-02": "Air Compressor No.2",
    "UL-TN-01": "Tank 75Q",
    "UL-TN-02": "Tank 300Q",
    "UL-AD-01": "Air Dryer",
    "UL-MDB-01": "Main Distribution Board No.1",
    "UL-MDB-02": "Main Distribution Board No.2",
    "UL-MDB-03": "Main Distribution Board No.3",
    "UL-MDB-04": "Main Distribution Board No.4",
    "UL-MDB-05": "Main Distribution Board No.5",
    "UL-SF-01": "Sand Filter Tank No.1",
    "UL-SF-02": "Sand Filter Tank No.2",
    "UL-CF-01": "Carbon Filter Tank No.1",
    "UL-CF-02": "Carbon Filter Tank No.2",
    "UL-RS-01": "Resin Tank No.1",
    "UL-RS-02": "Resin Tank No.2",
    "UL-DP-01": "Deep Well Pump No.1",
    "UL-DP-02": "Deep Well Pump No.2",
    "UL-TP-01": "Transfer Pump No.1",
    "UL-TP-02": "Transfer Pump No.2",
    "UL-TP-03": "Transfer Pump No.3",
    "UL-TP-04": "Transfer Pump No.4",
    "UL-FP-01": "Fire Pump System",
    "UL-RO-01": "Water Filter RO",
    "UL-AB-02": "Air Blower No.2",
    "UL-AB-03": "Air Blower No.3",
    "UL-AB-04": "Air Blower No.4",
    "UL-SP--01": "Submersible Pump 1",
    "UL-SP--02": "Submersible Pump 2",
    "UL-SP--03": "Submersible Pump 3",
    "UL-SP--04": "Submersible Pump 4",
    "UL-SP--05": "Submersible Pump 5",
    "UL-SP--06": "Submersible Pump 6",
    "UL-SP--07": "Submersible Pump 7",
    "UL-SP--08": "Submersible Pump 8",
    "UL-HD-01": "Hood 1",
    "UL-HD-02": "Hood 2",
    "UL-HD-03": "Hood 3",
    "UL-HD-04": "Hood 4",
    "UL-HD-05": "Hood 5",
    "UL-IN-01": "Intake No.1",
    "UL-IN-02": "Intake No.2",
    "UL-IN-03": "Intake No.3",
    "UL-IN-04": "Intake No.4",
    "UL-EX-01": "Exhaust No.1",
    "UL-EX-02": "Exhaust No.2",
    "UL-EX-03": "Exhaust No.3",
    "UL-EX-04": "Exhaust No.4",
    "UL-EX-05": "Exhaust No.5",
    "UL-EX-06": "Exhaust No.6",
    "UL-EX-07": "Exhaust No.7",
    "UL-EX-08": "Exhaust No.8",
    "UL-EX-09": "Exhaust No.9",
    "UL-TR-01": "Transformer No.1",
    "UL-TR-02": "Transformer No.2",
    "UL-TR-03": "Transformer No.3",
    "UL-TR-04": "Transformer No.4",
    "UL-TR-05": "Transformer No.5",
    "UL-UV-01": "UV Machine 1",
    "UL-UV-02": "UV Machine 2",
    "UL-UV-03": "UV Machine 3",
    "UL-LPG-01": "LPG Tank No.1",
    "UL-LPG-02": "LPG Tank No.2",
    "UL-LPG-03": "LPG Tank No.3",
    "UL-VP-01": "Vaporizer No.1",
    "UL-VP-02": "Vaporizer No.2",
    "UL-WH-01": "Washing Machine 1",
    "UL-WH-02": "Washing Machine 2",
    "UL-DR-01": "Dryer 1",
    "UL-DR-02": "Dryer 2",
    "UL-PP-01": "PP Filter No.01",
    "UL-PP-02": "PP Filter No.02",
    "UL-PP-03": "PP Filter No.03",
    "UL-PP-04": "PP Filter No.04",
    "UL-PP-05": "PP Filter No.05",
    "UL-PP-06": "PP Filter No.06",
    "UL-PP-07": "PP Filter No.07",
}

CANONICAL_ASSET_LOCATIONS = {
    "UL-HB-01": {"area": "Boiler Building", "detail": None},
    "UL-BL-01": {"area": "Boiler Building", "detail": None},
    "UL-BL-02": {"area": "Boiler Building", "detail": None},
    "UL-AC-01": {"area": "Air pump room", "detail": None},
    "UL-AC-02": {"area": "Air pump room", "detail": None},
    "UL-AD-01": {"area": "Air pump room", "detail": None},
    "UL-MDB-01": {"area": "Main Electrical Control Room", "detail": "Main Electrical Control Room 1"},
    "UL-MDB-02": {"area": "Main Electrical Control Room", "detail": "Main Electrical Control Room 2"},
    "UL-MDB-03": {"area": "Main Electrical Control Room", "detail": "Main Electrical Control Room 3"},
    "UL-MDB-04": {"area": "Main Electrical Control Room", "detail": "Main Electrical Control Room 4 (Wastewater Treatment Plant)"},
    "UL-MDB-05": {"area": "Main Electrical Control Room", "detail": "Main Electrical Control Room 5"},
    "UL-TN-01": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-TN-02": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-SF-01": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-SF-02": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-CF-01": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-CF-02": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-DP-01": {"area": "Water Treatment Plant", "detail": "Deep Well Pump 1"},
    "UL-DP-02": {"area": "Water Treatment Plant", "detail": "Deep Well Pump 2"},
    "UL-RS-01": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-RS-02": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-TP-01": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-TP-02": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-TP-03": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-TP-04": {"area": "Water Treatment Plant", "detail": "Sewage treatment plant"},
    "UL-FP-01": {"area": "Water Treatment Plant", "detail": "Fire pump control room"},
    "UL-RO-01": {"area": "Water Treatment Plant", "detail": "Production building Level 2"},
    "UL-AB-02": {"area": "Wastewater Treatment Plant", "detail": "Wastewater treatment plant"},
    "UL-AB-03": {"area": "Wastewater Treatment Plant", "detail": "Wastewater treatment plant"},
    "UL-AB-04": {"area": "Wastewater Treatment Plant", "detail": "Wastewater treatment plant"},
    "UL-SP--01": {"area": "Wastewater Treatment Plant", "detail": "Sump.1 Production building"},
    "UL-SP--02": {"area": "Wastewater Treatment Plant", "detail": "Sump.1 Production building"},
    "UL-SP--03": {"area": "Wastewater Treatment Plant", "detail": "Sump. 2 In the front office building"},
    "UL-SP--04": {"area": "Wastewater Treatment Plant", "detail": "Sump. 2 In the front office building"},
    "UL-SP--05": {"area": "Wastewater Treatment Plant", "detail": "Sump 3, rear office building"},
    "UL-SP--06": {"area": "Wastewater Treatment Plant", "detail": "Sump 3, rear office building"},
    "UL-SP--07": {"area": "Wastewater Treatment Plant", "detail": "EQ pond of the wastewater treatment plant"},
    "UL-SP--08": {"area": "Wastewater Treatment Plant", "detail": "EQ pond of the wastewater treatment plant"},
    "UL-HD-01": {"area": "Cooking Area", "detail": "Cooking Line-D Room Air Extractor"},
    "UL-HD-02": {"area": "Cooking Area", "detail": "Cooking Line-D Room Air Extractor"},
    "UL-HD-03": {"area": "Cooking Area", "detail": "Cooking Line-D Room Air Extractor"},
    "UL-HD-04": {"area": "Cooking Area", "detail": "Cooking Line-D Room Air Extractor"},
    "UL-HD-05": {"area": "Cooking Area", "detail": "Cooking Line-B Room Air Vent"},
    "UL-IN-01": {"area": "Cooking Area", "detail": "Aeration system, Cooking Line-C room"},
    "UL-IN-02": {"area": "Cooking Area", "detail": "Aeration system, Cooking Line-D room"},
    "UL-IN-03": {"area": "Cooking Area", "detail": "Aeration system"},
    "UL-IN-04": {"area": "Cooking Area", "detail": "Aeration system, Cooking Line-B room"},
    "UL-EX-01": {"area": "Cooking Area", "detail": "Air suction system, raw side washing room"},
    "UL-EX-02": {"area": "Cooking Area", "detail": "Air suction system for Forming Line-A area"},
    "UL-EX-03": {"area": "Cooking Area", "detail": "Air suction system for the Cooking Line-D area"},
    "UL-EX-04": {"area": "Cooking Area", "detail": "Air suction system, Fryer Line-C area"},
    "UL-EX-05": {"area": "Cooking Area", "detail": "Air suction system, raw side electrical room"},
    "UL-EX-06": {"area": "Cooking Area", "detail": "Air suction system for Packing Line-B room area"},
    "UL-EX-07": {"area": "Cooking Area", "detail": "Air suction system in the Catoning Line-C room"},
    "UL-EX-08": {"area": "Cooking Area", "detail": "Air suction system"},
    "UL-EX-09": {"area": "Cooking Area", "detail": "Air suction system for the Cooking Line-B area"},
    "UL-TR-01": {"area": "Cooking Area", "detail": None},
    "UL-TR-02": {"area": "Cooking Area", "detail": None},
    "UL-TR-03": {"area": "Cooking Area", "detail": None},
    "UL-TR-04": {"area": "Cooking Area", "detail": None},
    "UL-TR-05": {"area": "Cooking Area", "detail": None},
    "UL-LPG-01": {"area": "LPG Gas station", "detail": None},
    "UL-LPG-02": {"area": "LPG Gas station", "detail": None},
    "UL-LPG-03": {"area": "LPG Gas station", "detail": None},
    "UL-VP-01": {"area": "LPG Gas station", "detail": None},
    "UL-VP-02": {"area": "LPG Gas station", "detail": None},
    "UL-UV-01": {"area": "Raw side", "detail": None},
    "UL-UV-02": {"area": "Raw side", "detail": None},
    "UL-UV-03": {"area": "Raw side", "detail": None},
    "UL-WH-01": {"area": "Laundry room", "detail": None},
    "UL-WH-02": {"area": "Laundry room", "detail": None},
    "UL-DR-01": {"area": "Laundry room", "detail": None},
    "UL-DR-02": {"area": "Laundry room", "detail": None},
    "UL-PP-01": {"area": "Ceiling", "detail": None},
    "UL-PP-02": {"area": "Ceiling", "detail": None},
    "UL-PP-03": {"area": "Ceiling", "detail": None},
    "UL-PP-04": {"area": "Ceiling", "detail": None},
    "UL-PP-05": {"area": "Ceiling", "detail": None},
    "UL-PP-06": {"area": "Ceiling", "detail": None},
    "UL-PP-07": {"area": "Ceiling", "detail": None},
}

CANONICAL_EQUIPMENT_NAMES = {
    "PDM-B-01": "Vacuum Tumbler Small",
    "PDM-B-02": "Vacuum Tumbler Medium",
    "PDM-B-03": "Vacuum Tumbler Large",
    "PDM-B-06": "Tank Loader Set",
    "PDM-B-19": "Heat Seal Machine - Horizontal Conveyor",
}


def get_canonical_asset_name(asset_code: str, fallback_name: str | None = None) -> str | None:
    canonical_name = CANONICAL_ASSET_NAMES.get(clean_text(asset_code).upper())
    if canonical_name:
        return canonical_name
    return clean_text(fallback_name)


def get_canonical_equipment_name(asset_code: str, fallback_name: str | None = None) -> str | None:
    canonical_name = CANONICAL_EQUIPMENT_NAMES.get(clean_text(asset_code).upper())
    if canonical_name:
        return canonical_name
    return clean_text(fallback_name)


def pick_preferred_name(current_name: str | None, candidate_name: str | None) -> str | None:
    current = clean_text(current_name)
    candidate = clean_text(candidate_name)

    if not current:
        return candidate
    if not candidate:
        return current

    current_has_thai = contains_thai(current)
    candidate_has_thai = contains_thai(candidate)

    if current_has_thai and not candidate_has_thai:
        return candidate
    if not current_has_thai and candidate_has_thai:
        return current

    return current


def normalize_name_token(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (clean_text(value or "") or "").lower()).strip()


def is_production_critical_name(name: str | None) -> bool:
    normalized = normalize_name_token(name)
    if not normalized:
        return False
    return (
        "combi oven" in normalized
        or re.search(r"\bcombi\b", normalized) is not None
        or "brattpan" in normalized
    )


def derive_priority(*, is_overdue: bool, is_pending: bool, is_production_critical: bool) -> str:
    if is_production_critical and (is_overdue or is_pending):
        return "High"
    if is_overdue:
        return "Medium"
    return "Normal"


def get_canonical_asset_location(asset_code: str, fallback_location: str | None = None) -> dict[str, str | None]:
    canonical_location = CANONICAL_ASSET_LOCATIONS.get(clean_text(asset_code).upper())
    if canonical_location:
        return {
            "area": canonical_location.get("area"),
            "detail": canonical_location.get("detail"),
        }
    normalized = normalize_location_display(fallback_location)
    return {
        "area": normalized,
        "detail": None,
    }


def get_sheet_due_dates(year: int, month: int, target_week: str):
    month_matrix = calendar.monthcalendar(year, month)
    monday_days = [week[calendar.MONDAY] for week in month_matrix if week[calendar.MONDAY] != 0]
    if not monday_days:
        monday_days = [week[calendar.TUESDAY] for week in month_matrix if week[calendar.TUESDAY] != 0]

    if target_week == "every_week":
        return [date(year, month, day) for day in monday_days]

    week_index = WEEK_ORDER.get(target_week)
    if week_index is None or not monday_days:
        return []

    if week_index == -1:
        return [date(year, month, monday_days[-1])]

    if len(monday_days) < week_index:
        return []

    return [date(year, month, monday_days[week_index - 1])]


def get_month_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def get_quarter_range(reference_date: date):
    quarter_start_month = ((reference_date.month - 1) // 3) * 3 + 1
    quarter_end_month = quarter_start_month + 2
    return quarter_start_month, quarter_end_month


def get_week_of_month(target_date: date) -> int:
    month_matrix = calendar.monthcalendar(target_date.year, target_date.month)
    for index, week in enumerate(month_matrix, start=1):
        if target_date.day in week:
            return index
    return max(1, ((target_date.day - 1) // 7) + 1)


def format_week_ordinal(week_number: int) -> str:
    if week_number == 1:
        return "1st week"
    if week_number == 2:
        return "2nd week"
    if week_number == 3:
        return "3rd week"
    if week_number == 4:
        return "4th week"
    return f"Week {week_number}"


def format_occurrence_week_phrase(target_date: date) -> str:
    due_dates = get_sheet_due_dates(target_date.year, target_date.month, "every_week")
    if due_dates and target_date == due_dates[-1]:
        week_label = "Last week"
    else:
        week_label = format_week_ordinal(get_week_of_month(target_date))
    return f"{week_label} of {MONTH_LABELS[target_date.month - 1]}"


def copy_asset(asset):
    return copy.deepcopy(asset)


def load_utility_asset_source():
    signature = get_file_signature(MAINTENANCE_WORKBOOK_PATH)
    cached = _MAINTENANCE_CACHE.get("asset_source")
    if cached and cached["signature"] == signature:
        return cached["data"]

    data = {
        "source_path": str(MAINTENANCE_WORKBOOK_PATH),
        "last_synced": None,
        "assets": [],
        "errors": [],
    }

    if signature is None:
        data["errors"].append("Maintenance workbook unavailable")
        return data

    try:
        workbook = pd.ExcelFile(MAINTENANCE_WORKBOOK_PATH, engine="openpyxl")
        workbook_styles = openpyxl.load_workbook(MAINTENANCE_WORKBOOK_PATH, data_only=False, keep_links=False)
        merged_assets = {}

        for sheet_name in workbook.sheet_names:
            if not str(sheet_name).strip().upper().startswith("UL"):
                continue

            frame = pd.read_excel(MAINTENANCE_WORKBOOK_PATH, sheet_name=sheet_name, engine="openpyxl", header=None)
            if not worksheet_is_utility_schedule(frame):
                continue

            worksheet = workbook_styles[sheet_name] if sheet_name in workbook_styles.sheetnames else None
            inspection_map = extract_utility_inspection_templates(worksheet, frame, sheet_name) if worksheet is not None else {}

            for asset in extract_schedule_assets(frame, sheet_name):
                asset["inspection_templates"] = inspection_map.get(asset["asset_code"], [])
                code = asset["asset_code"]
                if code not in merged_assets:
                    merged_assets[code] = asset
                    continue

                existing = merged_assets[code]
                for key in ["asset_name", "location_raw", "location_display"]:
                    if not existing.get(key) and asset.get(key):
                        existing[key] = asset[key]
                existing_templates = {
                    (item.get("month"), item.get("target_week"))
                    for item in existing.get("inspection_templates", [])
                }
                for template in asset.get("inspection_templates", []):
                    key = (template.get("month"), template.get("target_week"))
                    if key not in existing_templates:
                        existing.setdefault("inspection_templates", []).append(template)
                        existing_templates.add(key)

        assets = []
        for code in sorted(merged_assets.keys()):
            asset = merged_assets[code]
            schedule = SCHEDULE_RULES.get(code, {})
            asset["asset_name"] = get_canonical_asset_name(code, asset.get("asset_name"))
            location_info = get_canonical_asset_location(code, asset.get("location_raw") or asset.get("location_display"))
            asset["location_display"] = location_info.get("area")
            asset["location_detail"] = location_info.get("detail")
            asset["category"] = schedule.get("category", "Others")
            asset["subcategory"] = schedule.get("subcategory", "General")
            assets.append(asset)

        data["assets"] = assets
        data["last_synced"] = datetime.fromtimestamp(MAINTENANCE_WORKBOOK_PATH.stat().st_mtime).isoformat()
    except Exception as exc:
        data["errors"].append(str(exc))

    _MAINTENANCE_CACHE["asset_source"] = {
        "signature": signature,
        "data": data,
    }
    return data


def build_occurrence(asset, scheduled_date: date, today: date):
    completion_tracking_available = False
    scheduled_week_end = scheduled_date + timedelta(days=6)
    completed_date = None
    inferred_done = not completion_tracking_available and scheduled_week_end < today
    is_done = completed_date is not None or inferred_done
    is_overdue = completion_tracking_available and not is_done and scheduled_week_end < today
    is_due_this_week = scheduled_date.isocalendar()[:2] == today.isocalendar()[:2]
    is_due_this_month = scheduled_date.year == today.year and scheduled_date.month == today.month
    is_upcoming_7_days = not is_done and today < scheduled_date <= today + timedelta(days=7)

    if is_done:
        status = "Done"
    elif is_overdue:
        status = "Overdue"
    elif scheduled_date >= today:
        status = "Pending"
    else:
        status = "Upcoming"

    production_critical = is_production_critical_name(asset.get("asset_name"))
    days_overdue = max((today - scheduled_week_end).days, 0) if is_overdue else 0
    priority = derive_priority(
        is_overdue=is_overdue,
        is_pending=(not is_done and not is_overdue),
        is_production_critical=production_critical,
    )
    inspection_template = next(
        (
            template
            for template in asset.get("inspection_templates", [])
            if template.get("month") == scheduled_date.month and template.get("target_week") == asset.get("target_week")
        ),
        None,
    )
    inspection_required = bool(inspection_template and inspection_template.get("inspection_required"))
    inspection_label = ADDITIONAL_STEPS_LABEL if inspection_required else None

    return {
        "asset_code": asset["asset_code"],
        "asset_name": asset["asset_name"],
        "category": asset["category"],
        "subcategory": asset["subcategory"],
        "location_raw": asset["location_raw"],
        "location_display": asset["location_display"] or asset["location_raw"] or "Unknown",
        "location_detail": asset.get("location_detail"),
        "frequency_type": asset["frequency_type"],
        "frequency_value": asset["frequency_value"],
        "start_month": asset["start_month"],
        "target_week": asset["target_week"],
        "due_months": asset["due_months"],
        "planned_occurrence": f"{MONTH_LABELS[scheduled_date.month - 1]} {WEEK_LABELS.get(asset['target_week'], asset['target_week'])}",
        "planned_month": scheduled_date.month,
        "planned_month_label": MONTH_LABELS[scheduled_date.month - 1],
        "planned_week": WEEK_LABELS.get(asset["target_week"], asset["target_week"]),
        "scheduled_week_label": f"Week {get_week_of_month(scheduled_date)}",
        "month_key": get_month_key(scheduled_date.year, scheduled_date.month),
        "scheduled_date": scheduled_date.isoformat(),
        "scheduled_date_label": scheduled_date.strftime("%d %b %Y"),
        "scheduled_week_end": scheduled_week_end.isoformat(),
        "next_due_date": scheduled_date.isoformat(),
        "next_due_label": scheduled_date.strftime("%d %b %Y"),
        "completed_date": completed_date.isoformat() if completed_date else None,
        "completed_date_label": completed_date.strftime("%d %b %Y") if completed_date else None,
        "completion_tracking_available": completion_tracking_available,
        "is_inferred_done": inferred_done,
        "status": status,
        "is_done": is_done,
        "is_pending": not is_done and not is_overdue,
        "is_overdue": is_overdue,
        "is_due_this_week": is_due_this_week,
        "is_due_this_month": is_due_this_month,
        "is_upcoming_7_days": is_upcoming_7_days,
        "is_production_critical": production_critical,
        "criticality": "Production Critical" if production_critical else "Standard",
        "priority": priority,
        "days_overdue": days_overdue,
        "assigned_technician": asset.get("assigned_technician") or "",
        "inspection_required": inspection_required,
        "inspection_label": inspection_label,
        "remarks": asset.get("remarks"),
        "source_sheet": asset.get("source_sheet"),
    }


def build_utility_dataset(year: int | None = None):
    today = datetime.now().date()
    active_year = int(year or today.year)
    cache_key = f"utility_dataset:{active_year}"
    source_data = load_utility_asset_source()
    signature = (get_file_signature(MAINTENANCE_WORKBOOK_PATH), active_year)
    cached = _MAINTENANCE_CACHE.get(cache_key)
    if cached and cached["signature"] == signature:
        return cached["data"]

    assets = []
    for asset in source_data["assets"]:
        schedule = SCHEDULE_RULES.get(asset["asset_code"])
        if not schedule:
            continue

        normalized = copy_asset(asset)
        normalized.update(schedule)
        assets.append(normalized)

    occurrences = []
    for asset in assets:
        for due_month in asset["due_months"]:
            for due_date in get_sheet_due_dates(active_year, due_month, asset["target_week"]):
                occurrences.append(build_occurrence(asset, due_date, today))

    occurrences.sort(key=lambda item: (item["scheduled_date"], item["asset_code"]))

    dataset = {
        "meta": {
            "source_path": source_data["source_path"],
            "last_synced": source_data["last_synced"],
            "year": active_year,
            "asset_count": len(assets),
            "scheduled_occurrence_count": len(occurrences),
            "errors": source_data["errors"],
        },
        "assets": assets,
        "occurrences": occurrences,
    }

    _MAINTENANCE_CACHE[cache_key] = {
        "signature": signature,
        "data": dataset,
    }
    return dataset


def parse_month_param(month_value: str | None, fallback_year: int):
    if not month_value:
        current = datetime.now().date()
        return current.year, current.month

    cleaned = str(month_value).strip()
    match = re.match(r"^(\d{4})-(\d{2})$", cleaned)
    if match:
        return int(match.group(1)), int(match.group(2))

    month_number = MONTH_NAME_TO_NUMBER.get(cleaned[:3].lower())
    if month_number:
        return fallback_year, month_number

    current = datetime.now().date()
    return current.year, current.month


def filter_occurrences(
    occurrences,
    *,
    month_key=None,
    status="all",
    category="all",
    location="all",
    inspection="all",
    asset_code="all",
    search="",
    sort="due_date_asc",
    priority="all",
    critical="all",
    week="all",
):
    search_value = clean_text(search)
    filtered = []

    for item in occurrences:
        if month_key and item["month_key"] != month_key:
            continue

        if category and category.lower() != "all" and item["category"] != category:
            continue

        if location and location.lower() != "all" and item["location_display"] != location:
            continue

        if inspection and inspection.lower() != "all":
            requires_inspection = bool(item.get("inspection_required"))
            if inspection.lower() == "inspection" and not requires_inspection:
                continue
            if inspection.lower() == "standard" and requires_inspection:
                continue

        if asset_code and asset_code.lower() != "all" and item["asset_code"] != asset_code:
            continue

        if priority and priority.lower() != "all" and clean_text(item.get("priority")) != clean_text(priority):
            continue

        if critical and critical.lower() != "all":
            wants_critical = critical.lower() in {"1", "true", "yes", "critical", "production_critical"}
            if bool(item.get("is_production_critical")) != wants_critical:
                continue

        if week and week.lower() != "all":
            week_value = clean_text(week)
            if week_value not in {
                clean_text(item.get("scheduled_week_label")),
                clean_text(format_occurrence_week_phrase(pd.to_datetime(item["scheduled_date"], errors="coerce").date()))
                if pd.notna(pd.to_datetime(item["scheduled_date"], errors="coerce"))
                else None,
            }:
                continue

        if status and status.lower() != "all":
            status_key = status.lower()
            if status_key == "pending" and not item["is_pending"]:
                continue
            if status_key == "done" and not item["is_done"]:
                continue
            if status_key == "overdue" and not item["is_overdue"]:
                continue
            if status_key == "due_this_week" and not item["is_due_this_week"]:
                continue
            if status_key == "due_this_month" and not item["is_due_this_month"]:
                continue
            if status_key == "upcoming" and not item["is_upcoming_7_days"]:
                continue

        if search_value:
            haystack = " ".join(
                [
                    item["asset_code"],
                    item["asset_name"],
                    item["category"],
                    item["subcategory"],
                    item["location_display"],
                ]
            ).lower()
            if search_value.lower() not in haystack:
                continue

        filtered.append(item)

    reverse = sort.endswith("_desc")
    if sort.startswith("machine"):
        filtered.sort(key=lambda item: (item["asset_code"], item["scheduled_date"]), reverse=reverse)
    elif sort.startswith("category"):
        filtered.sort(key=lambda item: (item["category"], item["location_display"], item["asset_name"]), reverse=reverse)
    else:
        filtered.sort(key=lambda item: (item["scheduled_date"], item["asset_code"]), reverse=reverse)

    return filtered


def find_next_due_occurrence(primary, asset_occurrences, today: date):
    next_due = next(
        (
            item for item in asset_occurrences
            if pd.to_datetime(item["scheduled_date"], errors="coerce").date() > today
        ),
        None,
    )
    if next_due:
        return next_due

    due_months = primary.get("due_months") or []
    target_week = primary.get("target_week")
    for year in [today.year, today.year + 1]:
        for due_month in due_months:
            for due_date in get_sheet_due_dates(year, due_month, target_week):
                if due_date > today:
                    return {
                        "scheduled_date": due_date.isoformat(),
                        "scheduled_date_label": due_date.strftime("%d %b %Y"),
                        "planned_month_label": MONTH_LABELS[due_date.month - 1],
                        "planned_week": WEEK_LABELS.get(target_week, target_week),
                    }

    return None


def build_asset_summary_rows(dataset, filtered_occurrences, *, sort="due_date_asc"):
    today = datetime.now().date()
    grouped_occurrences = {}
    for item in filtered_occurrences:
        grouped_occurrences.setdefault(item["asset_code"], []).append(item)

    all_occurrences_by_asset = {}
    for item in dataset["occurrences"]:
        all_occurrences_by_asset.setdefault(item["asset_code"], []).append(item)

    summary_rows = []
    for asset_code, matching_occurrences in grouped_occurrences.items():
        matching_occurrences = sorted(matching_occurrences, key=lambda item: (item["scheduled_date"], item["asset_code"]))
        asset_occurrences = sorted(all_occurrences_by_asset.get(asset_code, []), key=lambda item: (item["scheduled_date"], item["asset_code"]))
        primary = matching_occurrences[0]

        if any(item["is_overdue"] for item in matching_occurrences):
            status = "Overdue"
        elif any(item["is_pending"] for item in matching_occurrences):
            status = "Pending"
        elif any(item["is_done"] for item in matching_occurrences):
            status = "Done"
        else:
            status = primary.get("status", "Pending")

        done_occurrences = [
            item for item in asset_occurrences
            if item["is_done"]
        ]
        latest_done = done_occurrences[-1] if done_occurrences else None

        next_due = find_next_due_occurrence(primary, asset_occurrences, today)

        latest_done_label = "--"
        latest_done_week_phrase = None
        if latest_done:
            latest_done_date_obj = pd.to_datetime(latest_done["scheduled_date"], errors="coerce")
            if pd.notna(latest_done_date_obj):
                latest_done_week_phrase = format_occurrence_week_phrase(latest_done_date_obj.date())
                latest_done_label = latest_done_week_phrase

        next_due_label = "--"
        next_due_week_phrase = None
        if next_due:
            next_due_date_obj = pd.to_datetime(next_due["scheduled_date"], errors="coerce")
            if pd.notna(next_due_date_obj):
                next_due_week_phrase = format_occurrence_week_phrase(next_due_date_obj.date())
                next_due_label = next_due_week_phrase

        summary_rows.append(
            {
                "asset_code": primary["asset_code"],
                "asset_name": primary["asset_name"],
                "category": primary["category"],
                "subcategory": primary["subcategory"],
                "risk_level": primary.get("risk_level"),
                "location_display": primary["location_display"],
                "location_detail": primary.get("location_detail"),
                "frequency_type": primary["frequency_type"],
                "frequency_value": primary["frequency_value"],
                "target_week": primary["target_week"],
                "planned_week": primary["planned_week"],
                "frequency_label_primary": primary.get("frequency_label_primary"),
                "frequency_label_secondary": primary.get("frequency_label_secondary"),
                "status": status,
                "latest_done_label": latest_done_label,
                "latest_done_week": latest_done_week_phrase,
                "latest_done_month_label": latest_done["planned_month_label"] if latest_done else None,
                "latest_done_date": latest_done["scheduled_date"] if latest_done else None,
                "next_due_label": next_due_label,
                "next_due_week": next_due_week_phrase,
                "next_due_date_label": next_due["scheduled_date_label"] if next_due else None,
                "next_due_date": next_due["scheduled_date"] if next_due else None,
                "matching_occurrence_count": len(matching_occurrences),
                "days_overdue": max((item.get("days_overdue") or 0) for item in matching_occurrences) if matching_occurrences else 0,
                "is_production_critical": any(item.get("is_production_critical") for item in matching_occurrences),
                "criticality": "Production Critical" if any(item.get("is_production_critical") for item in matching_occurrences) else "Standard",
                "priority": derive_priority(
                    is_overdue=status == "Overdue",
                    is_pending=status == "Pending",
                    is_production_critical=any(item.get("is_production_critical") for item in matching_occurrences),
                ),
                "assigned_technician": primary.get("assigned_technician") or "",
                "inspection_required": any(item.get("inspection_required") for item in matching_occurrences),
                "inspection_label": next((item.get("inspection_label") for item in matching_occurrences if item.get("inspection_label")), None),
                "follow_up_status": "Pending Follow-Up" if any(item.get("inspection_required") and not item.get("is_done") for item in matching_occurrences) else ("Closed" if any(item.get("inspection_required") for item in matching_occurrences) else "-"),
            }
        )

    reverse = sort.endswith("_desc")
    if sort.startswith("machine"):
        summary_rows.sort(key=lambda item: (item["asset_code"], item["next_due_date"] or "9999-12-31"), reverse=reverse)
    elif sort.startswith("category"):
        summary_rows.sort(key=lambda item: (item["category"], item["location_display"], item["asset_name"]), reverse=reverse)
    else:
        summary_rows.sort(key=lambda item: (item["next_due_date"] or "9999-12-31", item["asset_code"]), reverse=reverse)

    return summary_rows


def build_equipment_attention_rows(month_items):
    asset_rows = build_asset_summary_rows(
        {"occurrences": month_items},
        [item for item in month_items if item.get("is_production_critical") and not item.get("is_done")],
        sort="due_date_asc",
    )
    asset_rows.sort(
        key=lambda row: (
            0 if row.get("priority") == "High" else 1,
            0 if row.get("status") == "Overdue" else 1,
            row.get("next_due_date") or "9999-12-31",
        )
    )
    return asset_rows[:6]


def build_equipment_top_risky_rows(month_items):
    asset_rows = build_asset_summary_rows({"occurrences": month_items}, month_items, sort="due_date_asc")
    asset_rows.sort(
        key=lambda row: (
            0 if row.get("is_production_critical") else 1,
            0 if row.get("status") == "Overdue" else 1,
            0 if row.get("priority") == "High" else 1,
            row.get("next_due_date") or "9999-12-31",
        )
    )
    return asset_rows[:5]


def build_summary_payload(year: int | None = None):
    dataset = build_utility_dataset(year)
    today = datetime.now().date()
    current_week = [item for item in dataset["occurrences"] if item["is_due_this_week"]]
    current_month = [item for item in dataset["occurrences"] if item["is_due_this_month"]]
    overdue = [item for item in dataset["occurrences"] if item["is_overdue"]]
    upcoming_7_days = [item for item in dataset["occurrences"] if item["is_upcoming_7_days"]]

    quarter_start, quarter_end = get_quarter_range(today)
    next_month_year = today.year + 1 if today.month == 12 else today.year
    next_month = 1 if today.month == 12 else today.month + 1

    tasks_this_quarter = [
        item
        for item in dataset["occurrences"]
        if item["planned_month"] >= quarter_start and item["planned_month"] <= quarter_end
    ]
    tasks_next_month = [
        item
        for item in dataset["occurrences"]
        if item["planned_month"] == next_month and item["month_key"] == get_month_key(next_month_year, next_month)
    ]

    category_load = Counter(item["category"] for item in dataset["occurrences"])
    most_maintenance_heavy = category_load.most_common(1)[0][0] if category_load else None

    def completion_rate(items):
        if not items:
            return 0.0
        done_count = sum(1 for item in items if item["is_done"])
        return round((done_count / len(items)) * 100.0, 1)

    monthly_load_distribution = []
    for month in range(1, 13):
        month_key = get_month_key(dataset["meta"]["year"], month)
        monthly_items = [item for item in dataset["occurrences"] if item["month_key"] == month_key]
        monthly_load_distribution.append(
            {
                "month_key": month_key,
                "label": MONTH_LABELS[month - 1],
                "total": len(monthly_items),
                "done": sum(1 for item in monthly_items if item["is_done"]),
                "pending": sum(1 for item in monthly_items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in monthly_items if item["is_overdue"]),
            }
        )

    category_breakdown = []
    for category in sorted({item["category"] for item in dataset["occurrences"]}):
        category_items = [item for item in dataset["occurrences"] if item["category"] == category]
        category_breakdown.append(
            {
                "category": category,
                "scheduled": len(category_items),
                "done": sum(1 for item in category_items if item["is_done"]),
                "pending": sum(1 for item in category_items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in category_items if item["is_overdue"]),
            }
        )

    location_breakdown = []
    for location in sorted({asset.get("location_display") or "Unknown" for asset in dataset["assets"]}):
        location_items = [item for item in dataset["occurrences"] if item["location_display"] == location]
        location_breakdown.append(
            {
                "location": location,
                "count": len(location_items),
            }
        )

    return {
        "meta": dataset["meta"],
        "summary": {
            "due_this_week": len(current_week),
            "due_this_month": len(current_month),
            "pending_this_week": sum(1 for item in current_week if not item["is_done"]),
            "done_this_week": sum(1 for item in current_week if item["is_done"]),
            "pending_this_month": sum(1 for item in current_month if item["is_pending"] and not item["is_overdue"]),
            "done_this_month": sum(1 for item in current_month if item["is_done"]),
            "completion_rate_week": completion_rate(current_week),
            "completion_rate_month": completion_rate(current_month),
            "overdue_count": len(overdue),
            "upcoming_next_7_days": len(upcoming_7_days),
            "total_utility_assets": dataset["meta"]["asset_count"],
            "tasks_this_quarter": len(tasks_this_quarter),
            "tasks_due_next_month": len(tasks_next_month),
            "most_maintenance_heavy_category": most_maintenance_heavy,
        },
        "charts": {
            "monthly_load_distribution": monthly_load_distribution,
            "category_breakdown": category_breakdown,
            "location_breakdown": location_breakdown,
        },
    }


def build_monthly_payload(month_value: str | None = None, year: int | None = None):
    dataset = build_utility_dataset(year)
    selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
    month_key = get_month_key(selected_year, selected_month)
    month_items = [item for item in dataset["occurrences"] if item["month_key"] == month_key]

    done_count = sum(1 for item in month_items if item["is_done"])
    overdue_count = sum(1 for item in month_items if item["is_overdue"])
    pending_count = len(month_items) - done_count - overdue_count

    category_groups = []
    for category in sorted({item["category"] for item in month_items}):
        items = [item for item in month_items if item["category"] == category]
        category_groups.append(
            {
                "category": category,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    location_groups = []
    for location in sorted({item["location_display"] or "Unknown" for item in month_items}):
        items = [item for item in month_items if (item["location_display"] or "Unknown") == location]
        location_groups.append(
            {
                "location": location,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    inspection_groups = []
    inspection_group_configs = [
        ("inspection", "With Additional Steps", [item for item in month_items if item.get("inspection_required")]),
        ("standard", "Normal Checklist", [item for item in month_items if not item.get("inspection_required")]),
    ]
    for inspection_value, inspection_label, items in inspection_group_configs:
        if not items:
            continue
        inspection_groups.append(
            {
                "inspection": inspection_value,
                "label": inspection_label,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    return {
        "meta": dataset["meta"],
        "selected_month": {
            "month_key": month_key,
            "year": selected_year,
            "month": selected_month,
            "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}",
        },
        "counts": {
            "done": done_count,
            "pending": pending_count,
            "overdue": overdue_count,
            "total": len(month_items),
        },
        "chart": {
            "labels": ["Done", "Pending", "Overdue"],
            "values": [done_count, pending_count, overdue_count],
        },
        "category_groups": category_groups,
        "location_groups": location_groups,
        "inspection_groups": inspection_groups,
    }


def build_list_payload(month_value=None, status="all", category="all", location="all", inspection="all", asset_code="all", search="", sort="due_date_asc", year=None, aggregate="occurrence"):
    dataset = build_utility_dataset(year)
    cleaned_month_value = clean_text(month_value)
    if cleaned_month_value and cleaned_month_value.lower() == "all":
        selected_year = dataset["meta"]["year"]
        selected_month = None
        month_key = None
        selected_month_payload = {
            "month_key": "all",
            "year": selected_year,
            "month": None,
            "label": f"All Months {selected_year}",
        }
    else:
        selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
        month_key = get_month_key(selected_year, selected_month)
        selected_month_payload = {
            "month_key": month_key,
            "year": selected_year,
            "month": selected_month,
            "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}",
        }

    rows = filter_occurrences(
        dataset["occurrences"],
        month_key=month_key,
        status=status,
        category=category,
        location=location,
        inspection=inspection,
        asset_code=asset_code,
        search=search,
        sort=sort,
    )

    if clean_text(aggregate).lower() == "asset":
        rows = build_asset_summary_rows(dataset, rows, sort=sort)

    return {
        "meta": dataset["meta"],
        "selected_month": selected_month_payload,
        "filters": {
            "status": status,
            "category": category,
            "location": location,
            "inspection": inspection,
            "asset_code": asset_code,
            "search": search,
            "sort": sort,
            "aggregate": aggregate,
        },
        "rows": rows,
    }


def build_timeline_payload(year: int | None = None, month_value: str | None = None):
    dataset = build_utility_dataset(year)
    selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
    selected_month_key = get_month_key(selected_year, selected_month)
    months = []
    for month in range(1, 13):
        month_key = get_month_key(dataset["meta"]["year"], month)
        month_items = [item for item in dataset["occurrences"] if item["month_key"] == month_key]
        months.append(
            {
                "month_key": month_key,
                "month": month,
                "label": MONTH_LABELS[month - 1],
                "total": len(month_items),
                "done": sum(1 for item in month_items if item["is_done"]),
                "pending": sum(1 for item in month_items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in month_items if item["is_overdue"]),
            }
        )

    selected_items = [item for item in dataset["occurrences"] if item["month_key"] == selected_month_key]
    week_groups = {}
    for item in selected_items:
        scheduled_date = pd.to_datetime(item["scheduled_date"], errors="coerce")
        if pd.isna(scheduled_date):
            continue
        week_number = get_week_of_month(scheduled_date.date())
        label = f"Week {week_number}"
        if label not in week_groups:
            week_groups[label] = {"label": label, "scheduled": 0, "completed": 0, "pending": 0, "overdue": 0}
        week_groups[label]["scheduled"] += 1
        if item["is_done"]:
            week_groups[label]["completed"] += 1
        elif item["is_overdue"]:
            week_groups[label]["overdue"] += 1
        else:
            week_groups[label]["pending"] += 1

    weekly_progress = [
        week_groups[label]
        for label in sorted(week_groups.keys(), key=lambda value: int(value.split()[-1]))
    ]

    return {
        "meta": dataset["meta"],
        "months": months,
        "selected_month": {
            "month_key": selected_month_key,
            "year": selected_year,
            "month": selected_month,
            "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}",
        },
        "weekly_progress": weekly_progress,
    }


def build_filter_payload(year: int | None = None):
    dataset = build_utility_dataset(year)
    year_value = dataset["meta"]["year"]
    months = [
        {
            "value": get_month_key(year_value, month),
            "label": f"{MONTH_LABELS[month - 1]} {year_value}",
        }
        for month in range(1, 13)
    ]

    return {
        "meta": dataset["meta"],
        "months": months,
        "categories": sorted({asset["category"] for asset in dataset["assets"]}),
        "locations": sorted({asset.get("location_display") or "Unknown" for asset in dataset["assets"]}),
        "inspection_options": [
            {"value": "all", "label": "All Maintenance"},
            {"value": "inspection", "label": "With Additional Steps"},
            {"value": "standard", "label": "Normal Checklist"},
        ],
        "asset_codes": sorted({asset["asset_code"] for asset in dataset["assets"]}),
        "status_options": [
            {"value": "all", "label": "All"},
            {"value": "pending", "label": "Pending"},
            {"value": "done", "label": "Done"},
            {"value": "overdue", "label": "Overdue"},
            {"value": "due_this_week", "label": "Due This Week"},
            {"value": "due_this_month", "label": "Due This Month"},
        ],
        "sort_options": [
            {"value": "due_date_asc", "label": "Due Date"},
            {"value": "machine_asc", "label": "Machine Code"},
            {"value": "category_asc", "label": "Category"},
        ],
    }


def get_maintenance_last_synced():
    return load_utility_asset_source().get("last_synced")


def load_equipment_asset_source():
    workbook_path = resolve_equipment_workbook_path()
    signature = get_file_signature(workbook_path)
    cache_key = "equipment_asset_source"
    cached = _MAINTENANCE_CACHE.get(cache_key)
    if cached and cached["signature"] == signature:
        return cached["data"]

    disk_cached = load_json_cache(EQUIPMENT_MAINTENANCE_CACHE_PATH)
    if (
        disk_cached
        and disk_cached.get("cache_version") == EQUIPMENT_MAINTENANCE_CACHE_VERSION
        and tuple(disk_cached.get("signature", ())) == signature
        and isinstance(disk_cached.get("data"), dict)
    ):
        data = disk_cached["data"]
        _MAINTENANCE_CACHE[cache_key] = {
            "signature": signature,
            "data": data,
        }
        return data

    data = {
        "source_path": str(workbook_path),
        "last_synced": None,
        "assets": [],
        "errors": [],
    }

    if signature is None:
        data["errors"].append("Equipment maintenance workbook unavailable")
        return data

    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=False, keep_links=False)
        merged_assets = {}

        workbook_year = datetime.fromtimestamp(workbook_path.stat().st_mtime).year
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if not worksheet_has_equipment_header(worksheet):
                continue

            for asset in extract_equipment_assets(worksheet, sheet_name, workbook_year):
                code = asset["asset_code"]
                if code not in merged_assets:
                    merged_assets[code] = asset
                    continue

                existing = merged_assets[code]
                existing["asset_name"] = pick_preferred_name(existing.get("asset_name"), asset.get("asset_name"))
                existing_templates = {
                    (item["month"], item["target_week"])
                    for item in existing.get("due_templates", [])
                }
                for template in asset.get("due_templates", []):
                    key = (template["month"], template["target_week"])
                    if key not in existing_templates:
                        existing.setdefault("due_templates", []).append(template)
                        existing_templates.add(key)

                for key in ["location_raw", "location_display", "category", "subcategory", "risk_level"]:
                    if not existing.get(key) and asset.get(key):
                        existing[key] = asset[key]

                existing_weeks = set(existing.get("scheduled_weeks", []))
                for week_number in asset.get("scheduled_weeks", []):
                    if week_number not in existing_weeks:
                        existing.setdefault("scheduled_weeks", []).append(week_number)
                        existing_weeks.add(week_number)

                existing_dates = set(existing.get("schedule_dates", []))
                for scheduled_date in asset.get("schedule_dates", []):
                    if scheduled_date not in existing_dates:
                        existing.setdefault("schedule_dates", []).append(scheduled_date)
                        existing_dates.add(scheduled_date)

        assets = []
        for code in sorted(merged_assets.keys()):
            asset = merged_assets[code]
            asset["asset_name"] = get_canonical_equipment_name(code, asset.get("asset_name")) or code
            if not asset.get("risk_level"):
                asset["risk_level"] = get_equipment_risk_level(asset.get("subcategory") or asset.get("location_raw"))
            asset["subcategory"] = classify_equipment_subcategory(
                asset.get("asset_name"),
                asset.get("location_raw") or asset.get("subcategory"),
                asset.get("risk_level"),
            )
            asset["category"] = derive_equipment_category(asset["subcategory"], asset.get("risk_level"))
            asset["location_display"] = asset["subcategory"]
            assets.append(asset)

        data["assets"] = assets
        data["last_synced"] = datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat()
    except Exception as exc:
        data["errors"].append(str(exc))

    _MAINTENANCE_CACHE[cache_key] = {
        "signature": signature,
        "data": data,
    }
    write_json_cache(
        EQUIPMENT_MAINTENANCE_CACHE_PATH,
        {
            "cache_version": EQUIPMENT_MAINTENANCE_CACHE_VERSION,
            "signature": list(signature) if signature else None,
            "data": data,
        },
    )
    return data


def build_equipment_occurrence(asset, template, today: date, year: int):
    scheduled_date = pd.to_datetime(template["scheduled_date"], errors="coerce")
    if pd.isna(scheduled_date):
        return []

    scheduled_date = scheduled_date.date()
    occurrence = build_occurrence(
        {
            **asset,
            "frequency_type": "scheduled",
            "frequency_value": len(asset.get("schedule_dates", [])) or 1,
            "start_month": scheduled_date.month,
            "target_week": "scheduled",
            "due_months": sorted({pd.to_datetime(value).month for value in asset.get("schedule_dates", [])}),
        },
        scheduled_date,
        today,
    )
    occurrence["risk_level"] = asset.get("risk_level")
    occurrence["frequency_label_primary"] = format_occurrence_week_phrase(scheduled_date).split(" of ")[0]
    occurrence["frequency_label_secondary"] = MONTH_LABELS[scheduled_date.month - 1]
    occurrence["inspection_required"] = bool(template.get("inspection_required"))
    occurrence["inspection_label"] = ADDITIONAL_STEPS_LABEL if occurrence["inspection_required"] else None
    return [occurrence]


def build_equipment_dataset(year: int | None = None):
    today = datetime.now().date()
    active_year = int(year or today.year)
    cache_key = f"equipment_dataset:{active_year}"
    source_data = load_equipment_asset_source()
    signature = (get_file_signature(resolve_equipment_workbook_path()), active_year)
    cached = _MAINTENANCE_CACHE.get(cache_key)
    if cached and cached["signature"] == signature:
        return cached["data"]

    assets = [copy_asset(asset) for asset in source_data["assets"]]
    occurrences = []
    for asset in assets:
        templates = asset.get("due_templates") or [{"scheduled_date": scheduled_date} for scheduled_date in asset.get("schedule_dates", [])]
        for template in templates:
            occurrences.extend(build_equipment_occurrence(asset, template, today, active_year))

    occurrences.sort(key=lambda item: (item["scheduled_date"], item["asset_code"]))
    dataset = {
        "meta": {
            "source_path": source_data["source_path"],
            "last_synced": source_data["last_synced"],
            "year": active_year,
            "asset_count": len(assets),
            "occurrence_count": len(occurrences),
            "domain": "equipment",
            "errors": source_data["errors"],
        },
        "assets": assets,
        "occurrences": occurrences,
    }
    _MAINTENANCE_CACHE[cache_key] = {
        "signature": signature,
        "data": dataset,
    }
    return dataset


def _build_summary_payload_from_dataset(dataset, *, asset_label="utility"):
    today = datetime.now().date()
    current_week = [item for item in dataset["occurrences"] if item["is_due_this_week"]]
    current_month = [item for item in dataset["occurrences"] if item["is_due_this_month"]]
    overdue = [item for item in dataset["occurrences"] if item["is_overdue"]]
    upcoming_7_days = [item for item in dataset["occurrences"] if item["is_upcoming_7_days"]]

    quarter_start, quarter_end = get_quarter_range(today)
    next_month_year = today.year + 1 if today.month == 12 else today.year
    next_month = 1 if today.month == 12 else today.month + 1

    tasks_this_quarter = [
        item
        for item in dataset["occurrences"]
        if quarter_start <= item["planned_month"] <= quarter_end
    ]
    tasks_next_month = [
        item
        for item in dataset["occurrences"]
        if item["planned_month"] == next_month and item["month_key"] == get_month_key(next_month_year, next_month)
    ]

    def completion_rate(items):
        if not items:
            return 0.0
        done_count = sum(1 for item in items if item["is_done"])
        return round((done_count / len(items)) * 100.0, 1)

    summary = {
        "due_this_week": len(current_week),
        "due_this_month": len(current_month),
        "pending_this_week": sum(1 for item in current_week if not item["is_done"]),
        "done_this_week": sum(1 for item in current_week if item["is_done"]),
        "pending_this_month": sum(1 for item in current_month if item["is_pending"] and not item["is_overdue"]),
        "done_this_month": sum(1 for item in current_month if item["is_done"]),
        "completion_rate_week": completion_rate(current_week),
        "completion_rate_month": completion_rate(current_month),
        "overdue_count": len(overdue),
        "upcoming_next_7_days": len(upcoming_7_days),
        f"total_{asset_label}_assets": dataset["meta"]["asset_count"],
        "tasks_this_quarter": len(tasks_this_quarter),
        "tasks_due_next_month": len(tasks_next_month),
    }

    if asset_label == "equipment":
        risk_counter = Counter(asset.get("category", "Medium Risk") for asset in dataset["assets"])
        summary["risk_breakdown"] = {
            "high": risk_counter.get("High Risk", 0),
            "medium": risk_counter.get("Medium Risk", 0),
            "low": risk_counter.get("Low Risk", 0),
        }

    return {"meta": dataset["meta"], "summary": summary}


def _build_monthly_payload_from_dataset(dataset, month_value: str | None = None):
    selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
    month_key = get_month_key(selected_year, selected_month)
    month_items = [item for item in dataset["occurrences"] if item["month_key"] == month_key]

    done_count = sum(1 for item in month_items if item["is_done"])
    overdue_count = sum(1 for item in month_items if item["is_overdue"])
    pending_count = len(month_items) - done_count - overdue_count

    category_groups = []
    for category in sorted({item["category"] for item in month_items}):
        items = [item for item in month_items if item["category"] == category]
        category_groups.append(
            {
                "category": category,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    location_groups = []
    for location in sorted({item["location_display"] or "Unknown" for item in month_items}):
        items = [item for item in month_items if (item["location_display"] or "Unknown") == location]
        location_groups.append(
            {
                "location": location,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    inspection_groups = []
    inspection_group_configs = [
        ("inspection", "With Additional Steps", [item for item in month_items if item.get("inspection_required")]),
        ("standard", "Normal Checklist", [item for item in month_items if not item.get("inspection_required")]),
    ]
    for inspection_value, inspection_label, items in inspection_group_configs:
        if not items:
            continue
        inspection_groups.append(
            {
                "inspection": inspection_value,
                "label": inspection_label,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    risk_groups = []
    for risk_level in ["High", "Medium", "Low"]:
        items = [item for item in month_items if item.get("risk_level") == risk_level]
        if not items:
            continue
        risk_groups.append(
            {
                "risk_level": risk_level,
                "count": len(items),
                "done": sum(1 for item in items if item["is_done"]),
                "pending": sum(1 for item in items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in items if item["is_overdue"]),
            }
        )

    return {
        "meta": dataset["meta"],
        "selected_month": {
            "month_key": month_key,
            "year": selected_year,
            "month": selected_month,
            "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}",
        },
        "counts": {
            "done": done_count,
            "pending": pending_count,
            "overdue": overdue_count,
            "total": len(month_items),
        },
        "chart": {
            "labels": ["Done", "Pending", "Overdue"],
            "values": [done_count, pending_count, overdue_count],
        },
        "category_groups": category_groups,
        "location_groups": location_groups,
        "inspection_groups": inspection_groups,
        "risk_groups": risk_groups,
    }


def _build_list_payload_from_dataset(
    dataset,
    month_value=None,
    status="all",
    category="all",
    location="all",
    inspection="all",
    asset_code="all",
    search="",
    sort="due_date_asc",
    aggregate="occurrence",
    priority="all",
    critical="all",
    week="all",
):
    cleaned_month_value = clean_text(month_value)
    if cleaned_month_value and cleaned_month_value.lower() == "all":
        selected_year = dataset["meta"]["year"]
        selected_month_payload = {
            "month_key": "all",
            "year": selected_year,
            "month": None,
            "label": f"All Months {selected_year}",
        }
        month_key = None
    else:
        selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
        month_key = get_month_key(selected_year, selected_month)
        selected_month_payload = {
            "month_key": month_key,
            "year": selected_year,
            "month": selected_month,
            "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}",
        }

    rows = filter_occurrences(
        dataset["occurrences"],
        month_key=month_key,
        status=status,
        category=category,
        location=location,
        inspection=inspection,
        asset_code=asset_code,
        search=search,
        sort=sort,
        priority=priority,
        critical=critical,
        week=week,
    )
    if clean_text(aggregate).lower() == "asset":
        rows = build_asset_summary_rows(dataset, rows, sort=sort)

    return {
        "meta": dataset["meta"],
        "selected_month": selected_month_payload,
        "filters": {
            "status": status,
            "category": category,
            "location": location,
            "inspection": inspection,
            "asset_code": asset_code,
            "search": search,
            "sort": sort,
            "aggregate": aggregate,
            "priority": priority,
            "critical": critical,
            "week": week,
        },
        "rows": rows,
    }


def _build_timeline_payload_from_dataset(dataset, year: int | None = None, month_value: str | None = None):
    selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
    selected_month_key = get_month_key(selected_year, selected_month)
    months = []
    for month in range(1, 13):
        month_key = get_month_key(dataset["meta"]["year"], month)
        month_items = [item for item in dataset["occurrences"] if item["month_key"] == month_key]
        months.append(
            {
                "month_key": month_key,
                "month": month,
                "label": MONTH_LABELS[month - 1],
                "total": len(month_items),
                "done": sum(1 for item in month_items if item["is_done"]),
                "pending": sum(1 for item in month_items if item["is_pending"] and not item["is_overdue"]),
                "overdue": sum(1 for item in month_items if item["is_overdue"]),
            }
        )

    selected_items = [item for item in dataset["occurrences"] if item["month_key"] == selected_month_key]
    week_groups = {}
    for item in selected_items:
        scheduled_date = pd.to_datetime(item["scheduled_date"], errors="coerce")
        if pd.isna(scheduled_date):
            continue
        week_number = get_week_of_month(scheduled_date.date())
        label = f"Week {week_number}"
        if label not in week_groups:
            week_groups[label] = {"label": label, "scheduled": 0, "completed": 0}
        week_groups[label]["scheduled"] += 1
        if item["is_done"]:
            week_groups[label]["completed"] += 1

    weekly_progress = [
        week_groups[label]
        for label in sorted(week_groups.keys(), key=lambda value: int(value.split()[-1]))
    ]

    return {
        "meta": dataset["meta"],
        "months": months,
        "selected_month": {
            "month_key": selected_month_key,
            "year": selected_year,
            "month": selected_month,
            "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}",
        },
        "weekly_progress": weekly_progress,
    }


def _build_filter_payload_from_dataset(dataset):
    year_value = dataset["meta"]["year"]
    months = [
        {
            "value": get_month_key(year_value, month),
            "label": f"{MONTH_LABELS[month - 1]} {year_value}",
        }
        for month in range(1, 13)
    ]
    payload = {
        "meta": dataset["meta"],
        "months": months,
        "categories": sorted({asset["category"] for asset in dataset["assets"]}),
        "locations": sorted({asset.get("location_display") or "Unknown" for asset in dataset["assets"]}),
        "inspection_options": [
            {"value": "all", "label": "All Maintenance"},
            {"value": "inspection", "label": "With Additional Steps"},
            {"value": "standard", "label": "Normal Checklist"},
        ],
        "asset_codes": sorted({asset["asset_code"] for asset in dataset["assets"]}),
        "status_options": [
            {"value": "all", "label": "All"},
            {"value": "pending", "label": "Pending"},
            {"value": "done", "label": "Done"},
            {"value": "overdue", "label": "Overdue"},
            {"value": "due_this_week", "label": "Due This Week"},
            {"value": "due_this_month", "label": "Due This Month"},
        ],
        "sort_options": [
            {"value": "due_date_asc", "label": "Due Date"},
            {"value": "machine_asc", "label": "Machine Code"},
            {"value": "category_asc", "label": "Category"},
        ],
    }
    if dataset["meta"].get("domain") == "equipment":
        payload["risk_levels"] = ["High", "Medium", "Low"]
    return payload


def build_equipment_summary_payload(year: int | None = None):
    return _build_summary_payload_from_dataset(build_equipment_dataset(year), asset_label="equipment")


def build_equipment_monthly_payload(month_value: str | None = None, year: int | None = None):
    dataset = build_equipment_dataset(year)
    payload = _build_monthly_payload_from_dataset(dataset, month_value)
    month_key = payload["selected_month"]["month_key"]
    month_items = [item for item in dataset["occurrences"] if item["month_key"] == month_key]
    open_items = [item for item in month_items if not item["is_done"]]
    high_priority_open = [item for item in open_items if item.get("priority") == "High"]
    critical_open = [item for item in open_items if item.get("is_production_critical")]
    payload["counts"]["completion_rate"] = round((payload["counts"]["done"] / payload["counts"]["total"]) * 100.0, 1) if payload["counts"]["total"] else 0.0
    payload["counts"]["high_priority_open"] = len(high_priority_open)
    payload["counts"]["production_critical_open"] = len(critical_open)
    payload["critical_attention"] = build_equipment_attention_rows(month_items)
    payload["top_risky_equipment"] = build_equipment_top_risky_rows(month_items)
    return payload


def build_equipment_list_payload(
    month_value=None,
    status="all",
    category="all",
    location="all",
    inspection="all",
    asset_code="all",
    search="",
    sort="due_date_asc",
    year=None,
    aggregate="occurrence",
    priority="all",
    critical="all",
    week="all",
):
    return _build_list_payload_from_dataset(
        build_equipment_dataset(year),
        month_value=month_value,
        status=status,
        category=category,
        location=location,
        inspection=inspection,
        asset_code=asset_code,
        search=search,
        sort=sort,
        aggregate=aggregate,
        priority=priority,
        critical=critical,
        week=week,
    )


def build_equipment_timeline_payload(year: int | None = None, month_value: str | None = None):
    return _build_timeline_payload_from_dataset(build_equipment_dataset(year), year, month_value)


def build_equipment_filter_payload(year: int | None = None):
    return _build_filter_payload_from_dataset(build_equipment_dataset(year))


def get_equipment_maintenance_last_synced():
    return load_equipment_asset_source().get("last_synced")


OVERVIEW_STATUS_OPTIONS = [
    {"value": "all", "label": "All"},
    {"value": "scheduled", "label": "Scheduled"},
    {"value": "completed", "label": "Completed"},
    {"value": "pending", "label": "Pending"},
]

OVERVIEW_SORT_OPTIONS = [
    {"value": "date_asc", "label": "Date"},
    {"value": "date_desc", "label": "Date (Newest)"},
    {"value": "name_asc", "label": "Name"},
    {"value": "category_asc", "label": "Category"},
    {"value": "status_asc", "label": "Status"},
]

OVERVIEW_INSPECTION_KEYWORDS = (
    "inspection",
    "inspect",
    "calibration",
    "supplier",
    "audit",
    "safety",
    "checklist",
)


def _parse_iso_timestamp(value):
    cleaned = clean_text(value)
    if not cleaned:
        return None
    try:
        return datetime.fromisoformat(cleaned)
    except ValueError:
        return None


def _build_overview_month_payloads(year_value: int):
    return [
        {"value": "all", "label": f"All Months {year_value}"}
    ] + [
        {"value": get_month_key(year_value, month), "label": f"{MONTH_LABELS[month - 1]} {year_value}"}
        for month in range(1, 13)
    ]


def _normalize_overview_status(item, today: date | None = None):
    reference_date = today or datetime.now().date()
    if item.get("is_done"):
        return "Completed"

    scheduled_date = pd.to_datetime(item.get("scheduled_date"), errors="coerce")
    if pd.notna(scheduled_date) and scheduled_date.date() >= reference_date:
        return "Scheduled"

    return "Pending"


def _requires_inspection(item):
    haystack = " ".join(
        str(value or "")
        for value in [
            item.get("asset_name"),
            item.get("category"),
            item.get("subcategory"),
            item.get("location_display"),
            item.get("location_detail"),
            item.get("remarks"),
            item.get("source_sheet"),
        ]
    ).lower()
    return any(keyword in haystack for keyword in OVERVIEW_INSPECTION_KEYWORDS)


def _derive_follow_up_status(item):
    normalized_status = item.get("normalized_status") or "Pending"
    inspection_required = bool(item.get("inspection_required"))

    if inspection_required and normalized_status != "Completed":
        return "Pending Follow-Up"
    if inspection_required and normalized_status == "Completed":
        return "Closed"
    if normalized_status == "Pending":
        return "Pending"
    if normalized_status == "Completed":
        return "Closed"
    return "-"


def _build_maintenance_overview_dataset(year: int | None = None):
    utility_dataset = build_utility_dataset(year)
    equipment_dataset = build_equipment_dataset(year)
    active_year = utility_dataset["meta"]["year"]
    today = datetime.now().date()

    occurrences = []
    seen_keys = set()
    for maintenance_category, dataset in [("Utility", utility_dataset), ("Equipment", equipment_dataset)]:
        for item in dataset["occurrences"]:
            dedupe_key = (maintenance_category, item.get("asset_code"), item.get("scheduled_date"))
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            occurrence = copy.deepcopy(item)
            occurrence["maintenance_category"] = maintenance_category
            occurrence["normalized_status"] = _normalize_overview_status(occurrence, today)
            occurrence["overview_inspection_required"] = False
            occurrence["overview_follow_up_status"] = "-"
            occurrence["preventive_maintenance_required"] = "Yes" if occurrence.get("scheduled_date") else "No"
            occurrence["date_provided_label"] = (
                occurrence.get("completed_date_label")
                or occurrence.get("scheduled_date_label")
                or "-"
            )
            occurrences.append(occurrence)

    occurrences.sort(
        key=lambda item: (
            item.get("scheduled_date") or "9999-12-31",
            item.get("maintenance_category") or "",
            item.get("asset_code") or "",
        )
    )

    synced_candidates = [
        _parse_iso_timestamp(utility_dataset["meta"].get("last_synced")),
        _parse_iso_timestamp(equipment_dataset["meta"].get("last_synced")),
    ]
    last_synced_values = [value for value in synced_candidates if value]

    return {
        "meta": {
            "year": active_year,
            "domain": "overview",
            "last_synced": max(last_synced_values).isoformat() if last_synced_values else None,
            "source_paths": [
                utility_dataset["meta"].get("source_path"),
                equipment_dataset["meta"].get("source_path"),
            ],
            "errors": [
                *utility_dataset["meta"].get("errors", []),
                *equipment_dataset["meta"].get("errors", []),
            ],
            "occurrence_count": len(occurrences),
        },
        "months": _build_overview_month_payloads(active_year),
        "occurrences": occurrences,
    }


def _filter_maintenance_overview_occurrences(
    occurrences,
    *,
    month_key=None,
    category="all",
    status="all",
    search="",
):
    search_value = clean_text(search)
    filtered = []
    for item in occurrences:
        if month_key and item.get("month_key") != month_key:
            continue
        if category and category.lower() != "all" and item.get("maintenance_category") != category:
            continue
        if status and status.lower() != "all" and clean_text(item.get("normalized_status")) != clean_text(status):
            continue
        if search_value:
            haystack = " ".join(
                [
                    str(item.get("asset_code") or ""),
                    str(item.get("asset_name") or ""),
                    str(item.get("maintenance_category") or ""),
                    str(item.get("assigned_technician") or ""),
                    str(item.get("remarks") or ""),
                    str(item.get("location_display") or ""),
                ]
            ).lower()
            if search_value.lower() not in haystack:
                continue
        filtered.append(item)
    return filtered


def _sort_maintenance_overview_rows(rows, sort="date_asc"):
    cleaned_sort = clean_text(sort) or "date_asc"
    reverse = cleaned_sort.endswith("_desc")

    if cleaned_sort.startswith("name"):
        rows.sort(key=lambda item: ((item.get("asset_name") or "").lower(), item.get("scheduled_date") or "9999-12-31"), reverse=reverse)
    elif cleaned_sort.startswith("category"):
        rows.sort(key=lambda item: ((item.get("maintenance_category") or "").lower(), (item.get("asset_name") or "").lower(), item.get("scheduled_date") or "9999-12-31"), reverse=reverse)
    elif cleaned_sort.startswith("status"):
        status_order = {"Scheduled": 0, "Pending": 1, "Completed": 2}
        rows.sort(
            key=lambda item: (
                status_order.get(item.get("normalized_status"), 99),
                item.get("scheduled_date") or "9999-12-31",
                (item.get("asset_name") or "").lower(),
            ),
            reverse=reverse,
        )
    else:
        rows.sort(key=lambda item: (item.get("scheduled_date") or "9999-12-31", (item.get("asset_name") or "").lower()), reverse=reverse)
    return rows


def build_maintenance_overview_payload(
    month_value=None,
    status="all",
    category="all",
    search="",
    sort="date_asc",
    year: int | None = None,
):
    dataset = _build_maintenance_overview_dataset(year)
    cleaned_month_value = clean_text(month_value)
    if cleaned_month_value and cleaned_month_value.lower() == "all":
        selected_month = {
            "month_key": "all",
            "year": dataset["meta"]["year"],
            "month": None,
            "label": f"All Months {dataset['meta']['year']}",
        }
        month_key = None
    else:
        selected_year, selected_month_number = parse_month_param(month_value, dataset["meta"]["year"])
        month_key = get_month_key(selected_year, selected_month_number)
        selected_month = {
            "month_key": month_key,
            "year": selected_year,
            "month": selected_month_number,
            "label": f"{MONTH_LABELS[selected_month_number - 1]} {selected_year}",
        }

    filtered_rows = _filter_maintenance_overview_occurrences(
        dataset["occurrences"],
        month_key=month_key,
        category=category,
        status=status,
        search=search,
    )
    sorted_rows = _sort_maintenance_overview_rows(filtered_rows[:], sort=sort)

    scheduled_tasks = len(sorted_rows)
    completed_tasks = sum(1 for row in sorted_rows if row.get("normalized_status") == "Completed")
    preventive_required = sum(1 for row in sorted_rows if row.get("normalized_status") == "Scheduled")
    pending_tasks = max(scheduled_tasks - completed_tasks, 0)
    inspection_required = sum(1 for row in sorted_rows if row.get("overview_inspection_required"))
    pending_follow_up = sum(1 for row in sorted_rows if row.get("overview_follow_up_status") == "Pending Follow-Up")

    rows = []
    for item in sorted_rows:
        rows.append(
            {
                "date": item.get("scheduled_date_label") or "-",
                "scheduled_date": item.get("scheduled_date"),
                "asset_code": item.get("asset_code") or "-",
                "asset_name": item.get("asset_name") or "-",
                "category": item.get("maintenance_category") or "-",
                "preventive_maintenance_required": item.get("preventive_maintenance_required") or "No",
                "status": item.get("normalized_status") or "Pending",
                "date_provided": item.get("date_provided_label") or "-",
                "person_in_charge": item.get("assigned_technician") or "-",
                "inspection_required": "-" if not item.get("overview_inspection_required") else "Yes",
                "follow_up_status": item.get("overview_follow_up_status") or "-",
            }
        )

    return {
        "meta": dataset["meta"],
        "selected_month": selected_month,
        "filters": {
            "month": selected_month["month_key"],
            "category": category,
            "status": status,
            "search": search,
            "sort": sort,
        },
        "filter_options": {
            "months": dataset["months"],
            "categories": [
                {"value": "all", "label": "All Categories"},
                {"value": "Utility", "label": "Utility"},
                {"value": "Equipment", "label": "Equipment"},
            ],
            "status_options": OVERVIEW_STATUS_OPTIONS,
            "sort_options": OVERVIEW_SORT_OPTIONS,
        },
        "summary": {
            "preventive_maintenance_required": preventive_required,
            "scheduled_tasks": scheduled_tasks,
            "completed_tasks": completed_tasks,
            "pending_tasks": pending_tasks,
            "completion_rate": round((completed_tasks / scheduled_tasks) * 100.0, 1) if scheduled_tasks else 0.0,
            "tasks_requiring_inspection": inspection_required,
            "tasks_pending_follow_up": pending_follow_up,
        },
        "rows": rows,
    }


NON_SCHEDULED_WORK_ORDERS = [
    {
        "work_order_id": "WO-260401",
        "maintenance_order_id": "MO-9001",
        "machine_code": "PDM-C-11",
        "machine_name": "Combi Oven 1",
        "area": "Cooking",
        "description": "Heating inconsistency reported during morning batch",
        "created_at": "2026-04-01T08:30:00",
        "due_at": "2026-04-02T17:00:00",
        "completed_at": None,
        "status": "Open",
        "priority": "",
        "technician": "",
        "downtime_hours": 1.5,
        "remarks": "Awaiting spare sensor confirmation",
        "source": "D365",
    },
    {
        "work_order_id": "WO-260402",
        "maintenance_order_id": "MO-9002",
        "machine_code": "PDM-BR-02",
        "machine_name": "Brattpan 2",
        "area": "Cooking",
        "description": "Steam jacket pressure check",
        "created_at": "2026-04-03T09:15:00",
        "due_at": "2026-04-06T16:00:00",
        "completed_at": None,
        "status": "Open",
        "priority": "",
        "technician": "",
        "downtime_hours": 0.5,
        "remarks": "Temporary manual workaround in place",
        "source": "D365",
    },
    {
        "work_order_id": "WO-260403",
        "maintenance_order_id": "MO-9003",
        "machine_code": "PDM-P-08",
        "machine_name": "Packing Conveyor 3",
        "area": "Packing",
        "description": "Belt tracking adjustment",
        "created_at": "2026-04-02T10:00:00",
        "due_at": "2026-04-04T12:00:00",
        "completed_at": None,
        "status": "Overdue",
        "priority": "",
        "technician": "Somchai",
        "downtime_hours": 2.0,
        "remarks": "",
        "source": "D365",
    },
    {
        "work_order_id": "WO-260404",
        "maintenance_order_id": "MO-9004",
        "machine_code": "PDM-O-03",
        "machine_name": "Outgoing Sealer 1",
        "area": "Out going",
        "description": "Seal jaw replacement",
        "created_at": "2026-04-04T13:20:00",
        "due_at": "2026-04-08T16:30:00",
        "completed_at": None,
        "status": "Open",
        "priority": "Medium",
        "technician": "Anan",
        "downtime_hours": 0.0,
        "remarks": "",
        "source": "D365",
    },
    {
        "work_order_id": "WO-260405",
        "maintenance_order_id": "MO-9005",
        "machine_code": "PDM-PR-07",
        "machine_name": "Vegetable Cutter 1",
        "area": "Preparation",
        "description": "Blade housing vibration",
        "created_at": "2026-04-01T14:00:00",
        "due_at": "2026-04-05T18:00:00",
        "completed_at": "2026-04-05T15:45:00",
        "status": "Closed",
        "priority": "",
        "technician": "Niran",
        "downtime_hours": 1.0,
        "remarks": "Completed and tested",
        "source": "D365",
    },
    {
        "work_order_id": "WO-260406",
        "maintenance_order_id": "MO-9006",
        "machine_code": "PDM-P-11",
        "machine_name": "Packing Checkweigher 2",
        "area": "Packing",
        "description": "Calibration drift",
        "created_at": "2026-04-06T09:50:00",
        "due_at": "2026-04-09T12:00:00",
        "completed_at": None,
        "status": "Open",
        "priority": "",
        "technician": "",
        "downtime_hours": 0.0,
        "remarks": "",
        "source": "D365",
    },
]


def build_non_scheduled_dataset(year: int | None = None):
    active_year = int(year or datetime.now().year)
    work_orders = []
    today = datetime.now()
    for item in NON_SCHEDULED_WORK_ORDERS:
        created_at = datetime.fromisoformat(item["created_at"])
        due_at = datetime.fromisoformat(item["due_at"])
        completed_at = datetime.fromisoformat(item["completed_at"]) if item.get("completed_at") else None
        production_critical = is_production_critical_name(item.get("machine_name"))
        normalized_status = clean_text(item.get("status")) or "Open"
        is_closed = normalized_status.lower() == "closed"
        is_overdue = not is_closed and due_at < today
        if is_overdue:
            normalized_status = "Overdue"
        priority = clean_text(item.get("priority")) or derive_priority(
            is_overdue=is_overdue,
            is_pending=not is_closed and not is_overdue,
            is_production_critical=production_critical,
        )
        work_orders.append(
            {
                **item,
                "created_at": created_at.isoformat(),
                "due_at": due_at.isoformat(),
                "completed_at": completed_at.isoformat() if completed_at else None,
                "status": normalized_status,
                "priority": priority,
                "criticality": "Production Critical" if production_critical else "Standard",
                "is_production_critical": production_critical,
                "days_overdue": max((today.date() - due_at.date()).days, 0) if is_overdue else 0,
                "month_key": get_month_key(due_at.year, due_at.month),
                "week_label": f"Week {get_week_of_month(due_at.date())}",
            }
        )
    return {
        "meta": {"year": active_year, "domain": "non_scheduled", "last_synced": datetime.now().isoformat()},
        "work_orders": work_orders,
    }


def build_non_scheduled_summary_payload(year: int | None = None):
    dataset = build_non_scheduled_dataset(year)
    orders = dataset["work_orders"]
    now = datetime.now()
    current_week = [item for item in orders if datetime.fromisoformat(item["due_at"]).isocalendar()[:2] == now.isocalendar()[:2]]
    current_month = [item for item in orders if item["month_key"] == get_month_key(now.year, now.month)]
    summary = {
        "open_work_orders": sum(1 for item in orders if item["status"] in {"Open", "Overdue"}),
        "closed_work_orders": sum(1 for item in orders if item["status"] == "Closed"),
        "overdue_work_orders": sum(1 for item in orders if item["status"] == "Overdue"),
        "high_priority_work_orders": sum(1 for item in orders if item["priority"] == "High"),
        "production_critical_open_work_orders": sum(1 for item in orders if item["is_production_critical"] and item["status"] in {"Open", "Overdue"}),
        "this_week_work_orders": len(current_week),
        "this_month_work_orders": len(current_month),
    }
    return {"meta": dataset["meta"], "summary": summary}


def build_non_scheduled_monthly_payload(month_value: str | None = None, year: int | None = None):
    dataset = build_non_scheduled_dataset(year)
    selected_year, selected_month = parse_month_param(month_value, dataset["meta"]["year"])
    month_key = get_month_key(selected_year, selected_month)
    rows = [item for item in dataset["work_orders"] if item["month_key"] == month_key]
    status_counts = Counter(item["status"] for item in rows)
    area_groups = []
    for area in sorted({item["area"] for item in rows}):
        items = [item for item in rows if item["area"] == area]
        area_groups.append({"area": area, "count": len(items), "overdue": sum(1 for item in items if item["status"] == "Overdue")})
    priority_groups = []
    for priority in ["High", "Medium", "Normal"]:
        items = [item for item in rows if item["priority"] == priority]
        if items:
            priority_groups.append({"priority": priority, "count": len(items)})
    critical_attention = [item for item in rows if item["is_production_critical"] and item["status"] in {"Open", "Overdue"}]
    return {
        "meta": dataset["meta"],
        "selected_month": {"month_key": month_key, "label": f"{MONTH_LABELS[selected_month - 1]} {selected_year}"},
        "counts": {
            "open": status_counts.get("Open", 0),
            "closed": status_counts.get("Closed", 0),
            "overdue": status_counts.get("Overdue", 0),
            "high_priority": sum(1 for item in rows if item["priority"] == "High"),
            "production_critical_open": sum(1 for item in rows if item["is_production_critical"] and item["status"] in {"Open", "Overdue"}),
            "total": len(rows),
        },
        "chart": {"labels": ["Open", "Closed", "Overdue"], "values": [status_counts.get("Open", 0), status_counts.get("Closed", 0), status_counts.get("Overdue", 0)]},
        "area_groups": area_groups,
        "priority_groups": priority_groups,
        "critical_attention": critical_attention[:6],
    }


def build_non_scheduled_list_payload(month_value=None, status="all", priority="all", area="all", search="", year=None, sort="due_date_asc"):
    dataset = build_non_scheduled_dataset(year)
    cleaned_month_value = clean_text(month_value)
    month_key = None if cleaned_month_value and cleaned_month_value.lower() == "all" else get_month_key(*parse_month_param(month_value, dataset["meta"]["year"]))
    rows = []
    for item in dataset["work_orders"]:
        if month_key and item["month_key"] != month_key:
            continue
        if status != "all" and item["status"].lower() != status.lower():
            continue
        if priority != "all" and item["priority"].lower() != priority.lower():
            continue
        if area != "all" and item["area"] != area:
            continue
        if search:
            haystack = " ".join([item["work_order_id"], item["machine_code"], item["machine_name"], item["description"], item["area"]]).lower()
            if clean_text(search).lower() not in haystack:
                continue
        rows.append(item)
    reverse = sort.endswith("_desc")
    rows.sort(key=lambda item: item["due_at"], reverse=reverse)
    return {"meta": dataset["meta"], "rows": rows}


def build_non_scheduled_filter_payload(year: int | None = None):
    dataset = build_non_scheduled_dataset(year)
    year_value = dataset["meta"]["year"]
    months = [{"value": "all", "label": f"All Months {year_value}"}] + [
        {"value": get_month_key(year_value, month), "label": f"{MONTH_LABELS[month - 1]} {year_value}"}
        for month in range(1, 13)
    ]
    return {
        "meta": dataset["meta"],
        "months": months,
        "status_options": [{"value": "all", "label": "All"}, {"value": "open", "label": "Open"}, {"value": "closed", "label": "Closed"}, {"value": "overdue", "label": "Overdue"}],
        "priority_options": [{"value": "all", "label": "All Priorities"}, {"value": "high", "label": "High"}, {"value": "medium", "label": "Medium"}, {"value": "normal", "label": "Normal"}],
        "areas": sorted({item["area"] for item in dataset["work_orders"]}),
    }
